import math
import pandas as pd
from datetime import time as dtime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

SQUARE_OFF_TIME = dtime(15, 10)

def _get_buy_value(high_price: float, is_reversal: bool = False) -> float:
    """Calculate BUY Value: ceil for decimal, direct for whole."""
    if is_reversal:
        # For reversals, use the actual high price to ensure trade execution
        return float(high_price)
    else:
        # For normal breakouts, use ceil logic
        if high_price == int(high_price):  # Already whole number
            return float(high_price)
        else:  # Decimal number
            return float(math.ceil(high_price))

def _get_sell_value(low_price: float, is_reversal: bool = False) -> float:
    """Calculate SELL Value: floor for decimal, direct for whole."""
    if is_reversal:
        # For reversals, use the actual low price to ensure trade execution
        return float(low_price)
    else:
        # For normal breakouts, use floor logic
        if low_price == int(low_price):  # Already whole number
            return float(low_price)
        else:  # Decimal number
            return float(math.floor(low_price))

def run_floating_band_strategy(df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    """
    COMPLETELY FIXED Floating Band Intraday Strategy Implementation
    
    Fixed Issues:
    1. Momentum tracking after trade execution
    2. GoingHigh/GoingDown detection logic
    3. Reference band updates
    4. Continuous tracking throughout session
    
    Strategy Logic:
    1. Initial setup with first 5-min candle (9:15-9:20)
    2. Track GoingHigh (Higher Highs) and GoingDown (Lower Lows) momentum
    3. UBStock/LBStock breakouts from reference bands
    4. Trade value refinement with ceil/floor logic
    5. Continuous momentum tracking after trade execution
    6. Direction change reversals
    7. Square-off at 15:10
    """
    req = ["time", "open", "high", "low", "close"]
    missing = [c for c in req if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    if df.empty:
        return df.assign(UB=pd.NA, LB=pd.NA, Range=pd.NA, Signal=""), []

    d = df.copy().reset_index(drop=True)
    
    # Initialize columns
    d["Range"] = 0.0
    d["UB"] = 0.0
    d["LB"] = 0.0
    d["Signal"] = ""
    if "volume" not in d.columns:
        d["volume"] = 0

    trades = []
    current_trade = None

    # Step 1: Initial Setup
    initial_high = float(d.loc[0, "high"])
    initial_low = float(d.loc[0, "low"])
    initial_range = initial_high - initial_low
    initial_ub = initial_high + initial_range
    initial_lb = initial_low - initial_range
    
    d.loc[0, "Range"] = initial_range
    d.loc[0, "UB"] = initial_ub
    d.loc[0, "LB"] = initial_lb
    d.loc[0, "Signal"] = "Initial"

    # Tracking variables
    session_highest = initial_high    # Session's highest high for GoingHigh
    session_lowest = initial_low      # Session's lowest low for GoingDown
    current_trend = None              # "UP" or "DOWN"
    current_direction = None          # "UP" or "DOWN" - overall market direction after initial signal
    last_signal_candle_idx = 0        # Last valid signal candle
    
    # Reference bands (for breakout detection)
    reference_ub = initial_ub
    reference_lb = initial_lb
    reference_candle_idx = 0
    
    # Breakout tracking
    ubstock_breakout_occurred = False
    lbstock_breakout_occurred = False
    
    # Trade values
    buy_value = None
    sell_value = None
    waiting_for_buystock = False
    waiting_for_sellstock = False

    print(f"Initial Setup: H={initial_high:.2f}, L={initial_low:.2f}")
    print(f"Initial Bands: UB={initial_ub:.2f}, LB={initial_lb:.2f}")

    # Process each candle
    for i in range(1, len(d)):
        current_time = d.loc[i, "time"]
        t = current_time.time() if hasattr(current_time, "time") else current_time
        high = float(d.loc[i, "high"])
        low = float(d.loc[i, "low"])
        close = float(d.loc[i, "close"])
        
        # Calculate current candle's bands
        current_range = high - low
        current_ub = high + current_range
        current_lb = low - current_range
        
        d.loc[i, "Range"] = current_range
        d.loc[i, "UB"] = current_ub
        d.loc[i, "LB"] = current_lb

        print(f"\n{t} | H:{high:.1f} L:{low:.1f} C:{close:.1f}")
        print(f"   Session tracking: HighestH={session_highest:.1f} LowestL={session_lowest:.1f}")
        print(f"   Reference: UB={reference_ub:.1f} LB={reference_lb:.1f}")
        if buy_value is not None:
            print(f"   BUY Value: {buy_value:.1f} (waiting: {waiting_for_buystock})")
        if sell_value is not None:
            print(f"   SELL Value: {sell_value:.1f} (waiting: {waiting_for_sellstock})")

        # EOD Square-off
        if t >= SQUARE_OFF_TIME and current_trade:
            exit_price = close
            pnl = (exit_price - current_trade["entry_price"]) if current_trade["side"] == "LONG" else (current_trade["entry_price"] - exit_price)
            trades.append({
                "entry_time": current_trade["entry_time"],
                "side": current_trade["side"], 
                "entry_price": current_trade["entry_price"],
                "exit_time": current_time,
                "exit_price": exit_price,
                "reason": "EOD",
                "pnl": pnl
            })
            current_trade = None
            d.loc[i, "Signal"] = "EOD_SQUAREOFF"
            print(f"EOD Square-off @ {exit_price} | P&L: {pnl:+.2f}")
            continue

        signal_assigned = False

        # STEP 2A: UBStock (First breakout above reference UB)
        # Only trigger if no LONG trade is currently open AND we're in DOWN direction
        if not ubstock_breakout_occurred and high > reference_ub and not signal_assigned:
            # Check if we already have a LONG position
            if current_trade and current_trade["side"] == "LONG":
                # Already have LONG position - no new breakout needed
                print(f"   UBStock: {high:.2f} > {reference_ub:.2f} (IGNORED - LONG already open)")
            # Check if we're in UP direction - only allow UBStock if we're in DOWN direction
            elif current_direction == "UP":
                print(f"   UBStock: {high:.2f} > {reference_ub:.2f} (IGNORED - UP direction, wait for LB breakdown)")
            else:
                # No LONG position and in DOWN direction - trigger breakout
                d.loc[i, "Signal"] = "UBStock"
                ubstock_breakout_occurred = True
                
                buy_value = _get_buy_value(high)
                waiting_for_buystock = True
                
                signal_assigned = True
                print(f"   UBStock: {high:.2f} > {reference_ub:.2f}")
                print(f"   BUY Value: {buy_value:.2f}")

        # STEP 2B: LBStock (First breakout below reference LB)
        # Only trigger if no SHORT trade is currently open AND we're in UP direction
        if not lbstock_breakout_occurred and low < reference_lb and not signal_assigned:
            # Check if we already have a SHORT position
            if current_trade and current_trade["side"] == "SHORT":
                # Already have SHORT position - no new breakout needed
                print(f"   LBStock: {low:.2f} < {reference_lb:.2f} (IGNORED - SHORT already open)")
            # Check if we're in DOWN direction - only allow LBStock if we're in UP direction
            elif current_direction == "DOWN":
                print(f"   LBStock: {low:.2f} < {reference_lb:.2f} (IGNORED - DOWN direction, wait for UB breakout)")
            else:
                # No SHORT position and in UP direction - trigger breakout
                d.loc[i, "Signal"] = "LBStock"
                lbstock_breakout_occurred = True
                
                sell_value = _get_sell_value(low)
                waiting_for_sellstock = True
                
                signal_assigned = True
                print(f"   LBStock: {low:.2f} < {reference_lb:.2f}")
                print(f"   SELL Value: {sell_value:.2f}")

        # STEP 3A: BUYStock (High >= BUY Value) - HIGHEST PRIORITY
        if waiting_for_buystock and buy_value is not None and high >= buy_value and not signal_assigned:
            entry_price = buy_value
            d.loc[i, "Signal"] = "BUYStock"
            
            # Close existing short
            if current_trade and current_trade["side"] == "SHORT":
                pnl = current_trade["entry_price"] - entry_price
                trades.append({
                    "entry_time": current_trade["entry_time"],
                    "side": current_trade["side"],
                    "entry_price": current_trade["entry_price"],
                    "exit_time": current_time,
                    "exit_price": entry_price,
                    "reason": "REVERSAL",
                    "pnl": pnl
                })
                print(f"     Closed SHORT @ {entry_price} | P&L: {pnl:+.2f}")
            
            # Open new long
            current_trade = {
                "entry_time": current_time,
                "side": "LONG",
                "entry_price": entry_price
            }
            
            current_trend = "UP"
            last_signal_candle_idx = i
            waiting_for_buystock = False  # CRITICAL: Reset waiting flag
            
            # Reset breakout flags to allow new breakouts in same direction
            ubstock_breakout_occurred = False
            lbstock_breakout_occurred = False
            
            # Reset session tracking to current candle to continue momentum detection from this point
            # This ensures momentum detection continues working after trade execution
            session_highest = high
            session_lowest = low
            
            signal_assigned = True
            print(f"   BUYStock LONG @ {entry_price} | Trend: UP")

        # STEP 3B: SELLStock (Low <= SELL Value) - HIGHEST PRIORITY
        if waiting_for_sellstock and sell_value is not None and low <= sell_value and not signal_assigned:
            entry_price = sell_value
            d.loc[i, "Signal"] = "SELLStock"
            
            # Close existing long
            if current_trade and current_trade["side"] == "LONG":
                pnl = entry_price - current_trade["entry_price"]
                trades.append({
                    "entry_time": current_trade["entry_time"],
                    "side": current_trade["side"],
                    "entry_price": current_trade["entry_price"],
                    "exit_time": current_time,
                    "exit_price": entry_price,
                    "reason": "REVERSAL",
                    "pnl": pnl
                })
                print(f"     Closed LONG @ {entry_price} | P&L: {pnl:+.2f}")
            
            # Open new short
            current_trade = {
                "entry_time": current_time,
                "side": "SHORT",
                "entry_price": entry_price
            }
            
            current_trend = "DOWN"
            last_signal_candle_idx = i
            waiting_for_sellstock = False  # CRITICAL: Reset waiting flag
            
            # Reset breakout flags to allow new breakouts in same direction
            ubstock_breakout_occurred = False
            lbstock_breakout_occurred = False
            
            # Reset session tracking to current candle to continue momentum detection from this point
            # This ensures momentum detection continues working after trade execution
            session_highest = high
            session_lowest = low
            
            signal_assigned = True
            print(f"   SELLStock SHORT @ {entry_price} | Trend: DOWN")

        # STEP 5: Direction Change Reversals
        if not signal_assigned and current_trend and last_signal_candle_idx < i:
            last_high = d.loc[last_signal_candle_idx, "high"]
            last_low = d.loc[last_signal_candle_idx, "low"]
            last_range = last_high - last_low
            last_ub = last_high + last_range
            last_lb = last_low - last_range
            
            # UP trend reversal
            if current_trend == "UP" and low < last_lb:
                reversal_sell_value = _get_sell_value(low, is_reversal=True)
                
                # Exit long first
                if current_trade and current_trade["side"] == "LONG":
                    pnl = close - current_trade["entry_price"]
                    trades.append({
                        "entry_time": current_trade["entry_time"],
                        "side": current_trade["side"],
                        "entry_price": current_trade["entry_price"],
                        "exit_time": current_time,
                        "exit_price": close,
                        "reason": "DIRECTION_CHANGE_EXIT",
                        "pnl": pnl
                    })
                    print(f"     Closed LONG @ {close} | P&L: {pnl:+.2f}")
                    current_trade = None
                
                # IMMEDIATE EXECUTION: Execute the reversal trade immediately
                current_trade = {
                    "entry_time": current_time,
                    "side": "SHORT",
                    "entry_price": reversal_sell_value
                }
                d.loc[i, "Signal"] = "SELLStock"
                print(f"   SELLStock SHORT @ {reversal_sell_value:.2f} | Trend: DOWN")
                
                # Reset waiting flags since trade is executed
                sell_value = None
                waiting_for_sellstock = False
                
                # Reset everything
                current_trend = "DOWN"
                last_signal_candle_idx = i
                reference_ub = current_ub
                reference_lb = current_lb
                reference_candle_idx = i
                # Update session tracking for new trend direction after reversal
                session_highest = high  # Reset to reversal candle high
                session_lowest = low    # Reset to reversal candle low
                ubstock_breakout_occurred = False
                lbstock_breakout_occurred = True
                buy_value = None
                waiting_for_buystock = False
                
                signal_assigned = True
                print(f"   Direction Change: Low {low:.2f} < last LB {last_lb:.2f}")
                
            # DOWN trend reversal
            elif current_trend == "DOWN" and high > last_ub:
                reversal_buy_value = _get_buy_value(high, is_reversal=True)
                
                # Exit short first
                if current_trade and current_trade["side"] == "SHORT":
                    pnl = current_trade["entry_price"] - close
                    trades.append({
                        "entry_time": current_trade["entry_time"],
                        "side": current_trade["side"],
                        "entry_price": current_trade["entry_price"],
                        "exit_time": current_time,
                        "exit_price": close,
                        "reason": "DIRECTION_CHANGE_EXIT",
                        "pnl": pnl
                    })
                    print(f"     Closed SHORT @ {close} | P&L: {pnl:+.2f}")
                    current_trade = None
                
                # IMMEDIATE EXECUTION: Execute the reversal trade immediately
                current_trade = {
                    "entry_time": current_time,
                    "side": "LONG",
                    "entry_price": reversal_buy_value
                }
                d.loc[i, "Signal"] = "BUYStock"
                print(f"   BUYStock LONG @ {reversal_buy_value:.2f} | Trend: UP")
                
                # Reset waiting flags since trade is executed
                buy_value = None
                waiting_for_buystock = False
                
                # Reset everything
                current_trend = "UP"
                last_signal_candle_idx = i
                reference_ub = current_ub
                reference_lb = current_lb
                reference_candle_idx = i
                # Update session tracking for new trend direction after reversal
                session_highest = high  # Reset to reversal candle high
                session_lowest = low    # Reset to reversal candle low
                ubstock_breakout_occurred = True
                lbstock_breakout_occurred = False
                sell_value = None
                waiting_for_sellstock = False
                
                signal_assigned = True
                print(f"   Direction Change: High {high:.2f} > last UB {last_ub:.2f}")

        # STEP 4: Direction Change Detection (Higher Priority than Momentum)
        # Check if price has broken the opposite direction's reference band
        if not signal_assigned and current_direction is not None:
            if current_direction == "UP" and low < reference_lb:
                # UP direction broken - change to DOWN
                current_direction = "DOWN"
                print(f"   Direction Change: UP → DOWN (Low {low:.2f} < LB {reference_lb:.2f})")
                # Reset breakout flags to allow new breakouts in opposite direction
                ubstock_breakout_occurred = False
                lbstock_breakout_occurred = False
                signal_assigned = True
            elif current_direction == "DOWN" and high > reference_ub:
                # DOWN direction broken - change to UP
                current_direction = "UP"
                print(f"   Direction Change: DOWN → UP (High {high:.2f} > UB {reference_ub:.2f})")
                # Reset breakout flags to allow new breakouts in opposite direction
                ubstock_breakout_occurred = False
                lbstock_breakout_occurred = False
                signal_assigned = True

        # STEP 5: Momentum Detection (GoingHigh/GoingDown) - LOWER PRIORITY
        # Only check if no trade execution signal was assigned
        if not signal_assigned:
            # STEP 5A: GoingHigh Detection - Only in UP direction or when direction not set
            if (current_direction is None or current_direction == "UP") and high > session_highest:
                d.loc[i, "Signal"] = "GoingHigh"
                session_highest = high  # Update session tracking
                session_lowest = low    # Reset session_lowest to current candle for future GoingDown detection
                
                # Set market direction to UP if this is the first momentum signal
                if current_direction is None:
                    current_direction = "UP"
                    print(f"   Market Direction Set: UP")
                # Don't change direction if already set - only continue in current direction
                
                # Update reference bands to this GoingHigh candle
                reference_ub = current_ub
                reference_lb = current_lb
                reference_candle_idx = i
                last_signal_candle_idx = i
                
                # Reset for new reference
                ubstock_breakout_occurred = False
                lbstock_breakout_occurred = False
                buy_value = None
                sell_value = None
                
                signal_assigned = True
                print(f"   GoingHigh: {high:.2f} > {session_highest - (high - session_highest):.2f} (previous highest)")
                print(f"   Updated Reference: UB={reference_ub:.2f}, LB={reference_lb:.2f}")
            # Ignore GoingHigh if in DOWN direction
            elif current_direction == "DOWN" and high > session_highest:
                print(f"   GoingHigh IGNORED (direction DOWN, wait for direction change)")

            # STEP 5B: GoingDown Detection - Only in DOWN direction or when direction not set
            elif (current_direction is None or current_direction == "DOWN") and low < session_lowest:
                d.loc[i, "Signal"] = "GoingDown"
                session_lowest = low   # Update session tracking
                session_highest = high # Reset session_highest to current candle for future GoingHigh detection
                
                # Set market direction to DOWN if this is the first momentum signal
                if current_direction is None:
                    current_direction = "DOWN"
                    print(f"   Market Direction Set: DOWN")
                # Don't change direction if already set - only continue in current direction
                
                # Update reference bands to this GoingDown candle
                reference_ub = current_ub
                reference_lb = current_lb
                reference_candle_idx = i
                last_signal_candle_idx = i
                
                # Reset for new reference
                ubstock_breakout_occurred = False
                lbstock_breakout_occurred = False
                buy_value = None
                sell_value = None
                
                signal_assigned = True
                print(f"   GoingDown: {low:.2f} < {session_lowest + (session_lowest - low):.2f} (previous lowest)")
                print(f"   Updated Reference: UB={reference_ub:.2f}, LB={reference_lb:.2f}")
            # Ignore GoingDown if in UP direction
            elif current_direction == "UP" and low < session_lowest:
                print(f"   GoingDown IGNORED (direction UP, wait for direction change)")

        # ALWAYS update session tracking (even when no signal is assigned)
        if high > session_highest:
            session_highest = high
        if low < session_lowest:
            session_lowest = low

        if not signal_assigned:
            print(f"   No signal")

    # Final square-off
    if current_trade:
        final_price = d.loc[len(d)-1, "close"]
        pnl = (final_price - current_trade["entry_price"]) if current_trade["side"] == "LONG" else (current_trade["entry_price"] - final_price)
        trades.append({
            "entry_time": current_trade["entry_time"],
            "side": current_trade["side"],
            "entry_price": current_trade["entry_price"],
            "exit_time": d.loc[len(d)-1, "time"],
            "exit_price": final_price,
            "reason": "EOD_FINAL",
            "pnl": pnl
        })
        print(f"\nFinal Square-off @ {final_price} | P&L: {pnl:+.2f}")

    return d, trades

def save_strategy_report(annotated_df, trades, date_str):
    """Save strategy results to Excel and CSV files."""
    outdir = "reports"
    os.makedirs(outdir, exist_ok=True)
    
    safe_date = date_str.replace("-", "_")
    xlsx_file = os.path.join(outdir, f"BNF_FloatingBand_{safe_date}.xlsx")
    csv_file = os.path.join(outdir, f"BNF_FloatingBand_{safe_date}.csv")

    required_cols = ["time", "volume", "Range", "high", "low", "UB", "LB", "Signal"]
    
    if "volume" not in annotated_df.columns:
        annotated_df["volume"] = 0
    
    output_df = annotated_df[required_cols].copy()
    # Fix column names to match requirements: Time | Volume | Range | High | Low | UB | LB | Signal
    output_df = output_df.rename(columns={
        "time": "Time", 
        "volume": "Volume",
        "high": "High",
        "low": "Low"
    })
    
    if "Time" in output_df.columns:
        try:
            output_df["Time"] = pd.to_datetime(output_df["Time"]).dt.strftime("%H:%M")
        except:
            output_df["Time"] = output_df["Time"].astype(str).str[:5]
    
    numerical_cols = ["Range", "High", "Low", "UB", "LB"]
    for col in numerical_cols:
        if col in output_df.columns:
            output_df[col] = output_df[col].round(2)
    
    try:
        output_df.to_excel(xlsx_file, index=False)
        output_df.to_csv(csv_file, index=False)
        apply_excel_formatting(xlsx_file)
        
        print(f"\nReports saved:")
        print(f"   Excel: {xlsx_file}")
        print(f"   CSV: {csv_file}")
        
    except Exception as e:
        print(f"Warning: Could not save reports - {e}")

def apply_excel_formatting(xlsx_path):
    """Apply color coding to Excel file."""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active

        header_row = {cell.value: cell.column for cell in ws[1]}
        signal_col = header_row.get("Signal")
        high_col = header_row.get("High")
        low_col = header_row.get("Low")

        if not (signal_col and high_col and low_col):
            return

        light_green = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        light_red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        dark_green = PatternFill(start_color="FF006400", end_color="FF006400", fill_type="solid")
        dark_red = PatternFill(start_color="FF8B0000", end_color="FF8B0000", fill_type="solid")
        white_font = Font(color="FFFFFFFF", bold=True)

        for row in range(2, ws.max_row + 1):
            signal = ws.cell(row=row, column=signal_col).value
            
            if signal in ["UBStock", "GoingHigh"]:
                ws.cell(row=row, column=high_col).fill = light_green
            elif signal in ["LBStock", "GoingDown"]:
                ws.cell(row=row, column=low_col).fill = light_red
            elif signal in ["BUYStock", "BUY"]:
                high_cell = ws.cell(row=row, column=high_col)
                high_cell.fill = dark_green
                high_cell.font = white_font
            elif signal in ["SELLStock", "SELL"]:
                low_cell = ws.cell(row=row, column=low_col)
                low_cell.fill = dark_red
                low_cell.font = white_font

        wb.save(xlsx_path)
        
    except Exception as e:
        print(f"Warning: Could not apply Excel formatting - {e}")

def print_strategy_summary(trades, date_str):
    """Print strategy performance summary."""
    print(f"\nStrategy Performance Summary for {date_str}")
    print("=" * 65)
    
    if not trades:
        print("Result: No trades executed")
        return
        
    total_pnl = sum(t['pnl'] for t in trades)
    winning_trades = [t for t in trades if t['pnl'] > 0]
    losing_trades = [t for t in trades if t['pnl'] <= 0]
    
    win_rate = (len(winning_trades) / len(trades)) * 100 if trades else 0
    max_profit = max([t['pnl'] for t in trades]) if trades else 0
    max_loss = min([t['pnl'] for t in trades]) if trades else 0
    
    print(f"Total Trades: {len(trades)}")
    print(f"Total P&L: {total_pnl:+.2f}")
    print(f"Win Rate: {win_rate:.1f}% ({len(winning_trades)}W / {len(losing_trades)}L)")
    if max_profit > 0:
        print(f"Best Trade: {max_profit:+.2f}")
    if max_loss < 0:
        print(f"Worst Trade: {max_loss:+.2f}")
    
    reason_counts = {}
    for trade in trades:
        reason = trade.get('reason', 'Unknown')
        reason_counts[reason] = reason_counts.get(reason, 0) + 1
    
    if reason_counts:
        print(f"\nExit Reason Breakdown:")
        for reason, count in reason_counts.items():
            print(f"   {reason}: {count}")
    
    print("=" * 65)